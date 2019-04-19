from src.common.BasePage import BasePage
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC
from collections import OrderedDict
from datetime import datetime
from config.baseconf import reportType
import time, os
from openpyxl.reader.excel import load_workbook
from config import baseconf
from src.common import opexcel


class ClinicalReport(LoginPage):

    def __init__(self):

        pass
    autocase = baseconf.autoCase_path
    book = load_workbook(autocase)
    txtHid = "$('#txtHospitalNumber')"
    hid = (By.ID,'txtHospitalNumber')  #病历号
    btnTest = (By.ID, 'btnTest')  # 校验
    btnMerger = (By.ID, 'btnMerger')  # 合并
    btnCode = (By.ID, 'btnCode')  # 代码化
    btnCollap = (By.CLASS_NAME, 'collapBtn')  # 返回上一页
    CheckReportTypeA = (By.ID, 'txtCheckReportTypeA')  # 报告类别A
    CheckReportTypeB = (By.ID, 'txtCheckReportTypeB')  # 报告类别B
    shouShuList = (By.ID, 'selShouShuList')  # 手术次数
    OperationList = (By.ID,'divOperationList')  #手术阶段（A段，B段，C段）
    tbodyReportList = (By.ID,'tbodyReportList')  #手术报告
    btnRes = (By.ID,'btnRes')    #报告撤销
    StateZhCom =(By.CLASS_NAME,'StateZhCom') #报告状态
    tbodyReportList = (By.ID,'tbodyReportList')

    def input_Hid(self,hid):
        '''输入病历号'''

        WebDriverWait(self.driver, 10).until(lambda the_driver: the_driver.find_element(*ClinicalReport.hid).is_displayed())
        self.driver.execute_script(ClinicalReport.txtHid+'.val("' + hid + '")')
        alert=BasePage.alert_close()
        if alert =='不存在该病历号':
            return u'没有此病人信息'
        self.driver.execute_script(ClinicalReport.txtHid+".blur()")
        BasePage.alert_close()

    def undo(self):
        self.driver.find_element_by_id('btnRes').click()
        BasePage.alert_close()
        time.sleep(1)

    def selectType(self,type):
        """ 报告类型列表中选择报告类型"""
        BasePage.wait_loading()
        self.driver.find_element(*ClinicalReport.CheckReportTypeA).click()
        time.sleep(1)
        clickNum = 0
        reporttype = self.driver.find_element_by_css_selector('div[data-text="' + reportType[type] + '"]')
        while not (reporttype.is_displayed()):
            if clickNum == 2:
                return "该报告类别没有找到"
            else:
                self.driver.find_element(*ClinicalReport.CheckReportTypeA).click()
                clickNum = clickNum + 1
        reporttype.click()  # 选择该报告类别
        BasePage.wait(self.driver.find_element(*ClinicalReport.tbodyReportList))

    def yearSelect(self,shiduan, value):
        """shiduan：a:年期内；b：测评段；c：孕药控；
            value：2；5；10；"""
        BasePage.alert_close()
        BasePage.wait(self.driver.find_element(*ClinicalReport.OperationList))
        # 读取随访阶段
        try:
            table = self.driver.find_element_by_xpath('//div[@id="divOperationList"]/table[last()]')
            BasePage.wait(table.find_element_by_xpath('tbody/tr[4]/td[1]'))
            table.find_element_by_xpath('tbody/tr[4]/td[1]').click()
            time.sleep(1)
            selshiduan = Select(table.find_element_by_name("selShiDuan"))
            selshiduan.select_by_value(shiduan)
            year = Select(table.find_element_by_name('selNianQi'))
            year.select_by_value(value)
            BasePage.alert_close()
        except Exception as e:
            print (e)

    def jiaoyan(self,reportType, Hid):
        """通用类型：Img131I，ImgA，ImgB"""

        try:
            BasePage.wait(*ClinicalReport.tbodyReportList)
        except Exception as e:
            print(e)
        BasePage.alert_close()
        tbodyReportList = self.driver.find_element(*ClinicalReport.tbodyReportList)
        reportList = tbodyReportList.find_elements_by_tag_name('tr')
        if (len(reportList) == 1):
            if (reportList[0].text == '没有报告信息.'):
                return Hid + u'无报告'
        for j in range(len(reportList)):  # 遍历报告份数
            BasePage.wait_loading()
            reportList = tbodyReportList.find_elements_by_tag_name('tr')
            if (len(reportList) <= j):
                return (Hid + u"报告日期超出范围")
            BasePage.wait(reportList[j])
            reportList[j].click()
            # 判断是否有alert
            BasePage.alert_close()
            className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
            while ('StateNoCom' != className and 'StateCFCom' != className):
                BasePage.wait_loading()
                ClinicalReport.undo()
                BasePage.alert_close()
                BasePage.wait_loading()
                reportList = tbodyReportList.find_elements_by_tag_name('tr')
                if (len(reportList) <= j):
                    return (Hid + u"报告日期超出范围")
                reportList[j].click()  # 焦点重新回到该报告上
                BasePage.alert_close()
                className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
            BasePage.wait_loading()
            yuanshiText = ClinicalReport.readyuanshiReport(reportType)
            jiaoyanText = ClinicalReport.readjiaoyanReport(reportType)
            zimu = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
            sheet = ClinicalReport.book[reportType]
            row = opexcel.max_row(reportType) + 1
            sheet['A' + str(row)] = Hid
            i = 0
            for t in yuanshiText:
                sheet[zimu[i] + str(row)] = yuanshiText[t]
                i = i + 1
            for s in jiaoyanText:
                sheet[zimu[i] + str(row)] = jiaoyanText[s]
                i = i + 1
            row = row + 1
            ClinicalReport.book.save(ClinicalReport.autocase)
            BasePage.screenPage(reportType)

    def undo(self):
        self.driver.find_element(*ClinicalReport.btnRes).click()
        BasePage.alert_close()

    def readyuanshiReport(self,reportType):
        checkResult = ''
        checkConclusion = ''
        if reportType in ['Bchao_pre', 'Bchao_fellowup','B_FNA']:
            checkResult = self.driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[1]/div[5]/div[2]').text
            checkConclusion = self.driver.find_element_by_xpath(
                '//div[@id="divUltrasonography"]/div[1]/div[6]/div[2]').text
        if reportType == 'ImgB':
            checkResult = self.driver.find_element_by_xpath('//div[@id="divImagingExamination"]/div[1]/div[5]/div[2]').text
            checkConclusion = self.driver.find_element_by_xpath(
                '//div[@id="divImagingExamination"]/div[1]/div[6]/div[2]').text
        if reportType == 'ImgA':
            checkResult = self.driver.find_element_by_xpath('//div[@id="divImagingExaminationB"]/div[1]/div[5]/div[2]').text
            checkConclusion = self.driver.find_element_by_xpath(
                '//div[@id="divImagingExaminationB"]/div[1]/div[6]/div[2]').text
        if reportType == 'Img131I':
            checkResult = self.driver.find_element_by_xpath(
                '//div[@id="divNuclideImgExamination"]/div[1]/div[5]/div[2]').text
            checkConclusion = self.driver.find_element_by_xpath(
                '//div[@id="divNuclideImgExamination"]/div[1]/div[6]/div[2]').text
        if reportType == 'Bfna_cell':
            checkResult = self.driver.find_element_by_xpath('//div[@id="divImagingExamination"]/div[1]/div[5]/div[2]').text
            checkConclusion = self.driver.find_element_by_xpath(
                '//div[@id="divImagingExamination"]/div[1]/div[6]/div[2]').text

        return {
            "checkResult": checkResult,
            "checkConclusion": checkConclusion,
        }

