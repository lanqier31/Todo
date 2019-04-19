from openpyxl.reader.excel import load_workbook
from config import baseconf

autocase = baseconf.autoCase_path
book = load_workbook(autocase)


def WriteExcel(result, locator, sheetname):
    sheet = book.get_sheet_by_name(sheetname)
    sheet[locator]= result


def max_row(sheetname):
    sheet = book[sheetname]
    return sheet.max_row


def All_content(sheetname):
    contents=[]
    sheet = book[sheetname]
    for row in sheet.rows:
        for cell in row:
            con = str(cell.value)
            contents.append(con)
    return contents[1:]

